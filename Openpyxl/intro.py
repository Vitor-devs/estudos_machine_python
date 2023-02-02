import openpyxl as xl
from openpyxl import Workbook
#Cria um workbook
wb = xl.Workbook()
#Cria uma worksheet e ativa
ws = wb.active

#Renomeia o titulo da worksheet para 'Minha planilha'
ws.title = 'Minha planilha'
ws.title
#Criando uma planilha em wb com o nome de teste na primeira posição
ws1 = wb.create_sheet('Teste',0)
wb.sheetnames

for worksheet in wb:
    print(worksheet)

ws.sheet_properties
ws.sheet_properties.tabColor = '1072BA'
wb.save('teste.xlsx')

#Acessando dados da planilha

#Carrendo a workbook
wb = xl.load_workbook(r'C:\Users\AMD\Documents\Estudos_Machine_Python\Openpyxl\teste.xlsx')
teste = wb['Teste']
teste['A1'] = 'Inserindo dados'
teste.cell(row=2, column=2, value= 1000)

for x in range(1,20):
    for y in range(1,20):
        teste.cell(row=x, column=y, value=f'Coordenada {x}{y}')

wb.save('teste.xlsx')
